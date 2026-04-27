"""
Generates the output Excel workbook from the list of scoring results.

Sheets produced:
  Results     — one row per company × topic, all fields
  Summary     — pivot table: average score per company (rows) × topic (columns)
  Quote Audit — subset focused on quote verification, highlights unverified rows
  Errors      — any companies/topics that failed with an error message

Cells are colour-coded for instant readability:
  Score column  → red (0) → orange (1) → yellow (2) → light green (3) → green (4)
  Confidence    → same colour scale mapped to Low/Medium/High
  quote_verified→ red cell if False
"""

import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Hex fill colours (no leading #) — rubric scores run 0 (worst) to 4 (best)
_SCORE_FILLS = {
    0: "FFD9D9",   # red        — No Disclosure
    1: "FFE0B2",   # orange     — Awareness
    2: "FFF9C4",   # yellow     — Developing
    3: "DCEDC8",   # light green — Advanced
    4: "C8E6C9",   # green      — Leading
}
_CONF_FILLS = {
    "Low":    "FFD9D9",
    "Medium": "FFF9C4",
    "High":   "C8E6C9",
}
_FAIL_FILL = "FFD9D9"   # red for unverified quotes / errors

_COLUMN_WIDTHS = {
    "company": 22, "topic": 20, "score": 8, "score_label": 16,
    "confidence": 13, "rationale": 55,
    "supporting_quote": 55, "page_reference": 14,
    "quote_verified": 15, "fuzzy_match_score": 16, "audit_note": 45,
    "error": 50,
}


def create_output_excel(results: list[dict]) -> bytes:
    """
    Build and return a formatted Excel workbook as raw bytes.
    Callers write the bytes to disk (see api.py:_run_pipeline) and serve
    the file via FastAPI's FileResponse.
    """
    df = pd.DataFrame(results)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # -- Results sheet
        df.to_excel(writer, index=False, sheet_name="Results")

        # -- Summary pivot
        if "score" in df.columns and not df["score"].isna().all():
            pivot = (
                df.dropna(subset=["score"])
                .pivot_table(index="company", columns="topic",
                             values="score", aggfunc="mean")
                .round(2)
            )
            pivot.to_excel(writer, sheet_name="Summary")

        # -- Quote Audit sheet
        audit_cols = [c for c in
                      ["company", "topic", "score", "score_label",
                       "supporting_quote", "quote_verified",
                       "fuzzy_match_score", "audit_note"]
                      if c in df.columns]
        df[audit_cols].to_excel(writer, index=False, sheet_name="Quote Audit")

        # -- Errors sheet (only if there are errors)
        if "error" in df.columns:
            err_df = df[df["error"].notna()][
                [c for c in ["company", "topic", "error"]
                 if c in df.columns]
            ]
            if not err_df.empty:
                err_df.to_excel(writer, index=False, sheet_name="Errors")

    # Reload workbook to apply visual formatting (openpyxl can't do this inside ExcelWriter)
    buf.seek(0)
    wb = load_workbook(buf)

    if "Results" in wb.sheetnames:
        _format_sheet(wb["Results"])

    if "Quote Audit" in wb.sheetnames:
        _highlight_unverified(wb["Quote Audit"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _solid(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _format_sheet(ws):
    """Colour-code scores and confidence on the Results sheet."""
    headers = [cell.value for cell in ws[1]]

    # Bold, wrapped header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    score_col = _col_index(headers, "score")
    conf_col  = _col_index(headers, "confidence")
    ver_col   = _col_index(headers, "quote_verified")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if score_col and row[score_col - 1].value in _SCORE_FILLS:
            row[score_col - 1].fill = _solid(_SCORE_FILLS[row[score_col - 1].value])

        if conf_col and row[conf_col - 1].value in _CONF_FILLS:
            row[conf_col - 1].fill = _solid(_CONF_FILLS[row[conf_col - 1].value])

        if ver_col and row[ver_col - 1].value is False:
            row[ver_col - 1].fill = _solid(_FAIL_FILL)

    # Column widths
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            _COLUMN_WIDTHS.get(header, 15)
        )


def _highlight_unverified(ws):
    """Red-fill entire rows where quote_verified is False."""
    headers = [cell.value for cell in ws[1]]
    ver_col = _col_index(headers, "quote_verified")
    if not ver_col:
        return
    for row in ws.iter_rows(min_row=2):
        if row[ver_col - 1].value is False:
            for cell in row:
                cell.fill = _solid(_FAIL_FILL)


def _col_index(headers: list, name: str):
    """Return 1-based column index or None if the column doesn't exist."""
    try:
        return headers.index(name) + 1
    except ValueError:
        return None
